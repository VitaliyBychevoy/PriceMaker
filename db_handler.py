import openpyxl

wb = openpyxl.reader.excel.load_workbook(filename='DB.xlsx', data_only=True)


class DB_handler:

    #List all type machine (Trumpf, Thick Turret, Thin Turret)
    @staticmethod
    def get_list_type_machine() -> list:
        wb.active = 0
        sheet = wb.active
        list_type_machine = []
        for item in range(2, 5):
            list_type_machine.append(sheet['A' + str(item)].value)
        return list_type_machine

    #List all system of each type machine
    @staticmethod
    def get_list_type_system(type_machine: str=None):
        wb.active = 0
        sheet = wb.active
        list_type_system = []
        if type_machine is None:
            return "Оберіть тип машини."
        elif type_machine == 'Trumpf':
            for item in range(2, 9):
                list_type_system.append(sheet['C' + str(item)].value)
        elif type_machine == 'Thick Turret':
            for item in range(2, 21):
                list_type_system.append(sheet['E' + str(item)].value)
        elif type_machine == 'Thin Turret':
            for item in range(2, 5):
                list_type_system.append(sheet['G' + str(item)].value)            
        else:
            return 'Оберіть систему машини зі списка.'
        return list_type_system

    #Trumpf standard
    @staticmethod
    def trumpf_standard_get_list_punch_name_ua() -> list:
        wb.active = 4
        sheet = wb.active
        list_punch_name_ua = []
        for item in range(2, 9):
            list_punch_name_ua.append(sheet['I' + str(item)].value)
        return list_punch_name_ua

    @staticmethod
    def trumpf_standard_get_punch_name_en_from_ua(punch_name_ua: str) -> str:
        punch_name_en: str = None
        wb.active = 4
        sheet = wb.active
        for item in range(2, 8):
            if punch_name_ua == sheet['I' + str(item)].value:
                punch_name_en = sheet['H' + str(item)].value
                return punch_name_en
        return punch_name_en

    @staticmethod
    def trumpf_standard_get_station(diametr: float) -> str:
        range_diametr_station = []
        station: str = None

        wb.active = 4
        sheet = wb.active

        range_diametr_station = []
        station: str = None
        for item in range(3, 10):
            print(sheet['S' + str(item)].value)
            range_diametr_station = sheet['S' + str(item)].value.split('~')
            print(range_diametr_station)
            if item == 3 and diametr < float(range_diametr_station[0]):
                print(f"To small diametr of circumscribed_circle. It's less than {range_diametr_station[0]}")
                return '0.0'
            elif float(range_diametr_station[1]) < diametr:
                range_diametr_station = []
                continue 
            elif float(range_diametr_station[0]) < diametr <= float(range_diametr_station[1]):
                station = sheet['R' + str(item)].value
                return station
            # else float(range_diametr_station[1]) < diametr:
            else:
                print(f"To big diametr of circumscribed_circle. It's more than {range_diametr_station[1]}")
                return '80.0'


# wb.active = 0
# sheet = wb.active
# list_type_machine = []
# for item in range(2, 5):
#     list_type_machine.append(sheet['A' + str(item)].value)

# print(list_type_machine)


# for item in range(len(wb.sheetnames)):
#     print(item, f"-{wb.sheetnames[item]}")
# print("\n")

o1 = DB_handler()
print(o1.trumpf_standard_get_station(45))

